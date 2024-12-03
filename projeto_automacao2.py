'''Essa automação pega dados de um site específico e joga no excel Microsoft 365, criando
uma tabela com um gráfico, facilitando a compreesão administrativa
Foi projetada para um site especifico porém é possível criar para todas os sites de forma ampla
'''








import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import BarChart, Reference

# URL do site com a tabela de honorários
url = 'https://sinaep.org.br/tabela-de-honorarios/'

# Enviar requisição GET para o site
response = requests.get(url)

# Verificar se a requisição foi bem-sucedida
if response.status_code == 200:
    # Parse da página HTML com BeautifulSoup
    soup = BeautifulSoup(response.text, 'html.parser')

    # Encontrar a tabela na página
    tabela = soup.find('table')

    # Verificar se a tabela foi encontrada
    if not tabela:
        print("Tabela não encontrada na página. Verifique o seletor.")
        exit()  # Termina o script se a tabela não for encontrada

    # Coletar todas as linhas da tabela
    rows = tabela.find_all('tr')

    # Lista para armazenar os dados extraídos
    dados = []
    
    for row in rows:
        colunas = row.find_all(['td', 'th'])
        
        if colunas:
            dados.append([col.text.strip().replace('\r', '').replace('\n', '') for col in colunas])

    # Criar um DataFrame
    df = pd.DataFrame(dados[1:], columns=dados[0])

    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabela de Honorários"

    # Formatação das colunas
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Escrever o cabeçalho
    for col_num, header in enumerate(dados[0], start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Escrever os dados
    for row_num, row in enumerate(dados[1:], start=2):
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Ajustar largura das colunas com limite máximo
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        col_letter = col[0].column_letter
        # Define uma largura máxima de 20 caracteres
        ws.column_dimensions[col_letter].width = min(max_length + 2, 20)

    # Criar o gráfico
    chart = BarChart()
    chart.title = "Gráfico de Honorários"
    chart.x_axis.title = "Categorias"
    chart.y_axis.title = "Valores"

    # Referenciar dados e categorias
    data_range = Reference(ws, min_col=2, min_row=1, max_col=len(dados[0]), max_row=len(dados))
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(dados))

    # Adicionar dados ao gráfico
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(cats)

    # Adicionar gráfico à planilha
    ws.add_chart(chart, "G5")

    # Salvar o arquivo Excel
    wb.save("tabela_honorarios_personalizada.xlsx")
    print("Tabela criada com sucesso em 'tabela_honorarios_personalizada.xlsx'!")
else:
    print("Erro ao acessar o site:", response.status_code)


